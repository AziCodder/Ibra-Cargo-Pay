"""
Сервис экспорта проекта в Excel (xlsx).

Формирует файл с тремя листами:
- «Позиции» — номенклатура проекта (для admin включает cost_price и profit)
- «Заявки на оплату» — список заявок с остатками
- «Сводка по валютам» — итоговая финансовая сводка (total, paid, remaining, profit)
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.payment import Payment
from app.models.payment_request import PaymentRequest
from app.models.payment_request_item import PaymentRequestItem
from app.models.project import Project
from app.models.project_item import ProjectItem
from app.models.supplier import Supplier


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def _apply_header_style(ws, row: int, column_count: int) -> None:
    for col in range(1, column_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT


def _auto_size(ws, column_count: int) -> None:
    for col_idx in range(1, column_count + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


async def generate_project_excel(
    db: AsyncSession, project_id: int, *, is_admin: bool
) -> bytes:
    """Возвращает xlsx-файл с данными проекта в bytes."""

    # ── Проект ──
    project_q = (
        select(Project)
        .where(Project.id == project_id)
        .options(selectinload(Project.client))
    )
    project = (await db.execute(project_q)).scalar_one_or_none()
    if project is None:
        raise ValueError(f"Проект {project_id} не найден")

    # ── Позиции номенклатуры ──
    items_q = (
        select(ProjectItem, Supplier.full_name.label("supplier_name"))
        .outerjoin(Supplier, ProjectItem.supplier_id == Supplier.id)
        .where(ProjectItem.project_id == project_id)
        .order_by(ProjectItem.id)
    )
    items_rows = (await db.execute(items_q)).all()

    # ── Заявки на оплату с суммами платежей ──
    requests_q = (
        select(PaymentRequest)
        .where(PaymentRequest.project_id == project_id)
        .options(
            selectinload(PaymentRequest.items).selectinload(PaymentRequestItem.project_item),
            selectinload(PaymentRequest.payments),
        )
        .order_by(PaymentRequest.created_at.desc())
    )
    requests = (await db.execute(requests_q)).scalars().all()

    # ── Агрегация по валютам ──
    # total = price * quantity (без комиссии)
    # commission = price * commission% * quantity (отдельный расчёт)
    # profit = (price - COALESCE(cost_price, price)) * quantity
    effective_cost = func.coalesce(ProjectItem.cost_price, ProjectItem.price)
    summary_items_q = (
        select(
            ProjectItem.currency,
            func.sum(ProjectItem.price * ProjectItem.quantity).label("total"),
            func.sum(
                ProjectItem.price * (ProjectItem.commission / 100) * ProjectItem.quantity
            ).label("commission"),
            func.sum(
                (ProjectItem.price - effective_cost) * ProjectItem.quantity
            ).label("profit"),
        )
        .where(ProjectItem.project_id == project_id)
        .group_by(ProjectItem.currency)
    )
    summary_items = {
        row.currency: (
            Decimal(str(row.total or 0)),
            Decimal(str(row.commission or 0)),
            Decimal(str(row.profit or 0)),
        )
        for row in (await db.execute(summary_items_q)).all()
    }

    # Сумма учитывает только confirmed платежи
    summary_paid_q = (
        select(Payment.currency, func.sum(Payment.amount).label("paid"))
        .join(PaymentRequest, Payment.payment_request_id == PaymentRequest.id)
        .where(
            PaymentRequest.project_id == project_id,
            Payment.status == "confirmed",
        )
        .group_by(Payment.currency)
    )
    summary_paid = {
        row.currency: Decimal(str(row.paid or 0))
        for row in (await db.execute(summary_paid_q)).all()
    }

    # ── Создаём workbook ──
    wb = Workbook()
    assert wb.active is not None  # for type checkers

    # ——— Лист «Позиции» ———
    ws_items = wb.active
    ws_items.title = "Позиции"

    if is_admin:
        headers = [
            "№",
            "Наименование",
            "Детали",
            "Количество",
            "Поставщик",
            "Цена",
            "Себестоимость",
            "Комиссия, %",
            "Валюта",
            "Сумма",
            "Сумма с комиссией",
            "Прибыль",
        ]
    else:
        headers = [
            "№",
            "Наименование",
            "Детали",
            "Количество",
            "Поставщик",
            "Цена",
            "Комиссия, %",
            "Валюта",
            "Сумма",
            "Сумма с комиссией",
        ]
    ws_items.append(headers)
    _apply_header_style(ws_items, 1, len(headers))

    for idx, (item, supplier_name) in enumerate(items_rows, start=1):
        line_total = item.price * item.quantity
        effective_price = item.price * (Decimal("1") + item.commission / Decimal("100"))
        line_total_with_commission = effective_price * item.quantity
        if is_admin:
            effective_cost = item.cost_price if item.cost_price is not None else item.price
            profit = (item.price - effective_cost) * item.quantity
            ws_items.append(
                [
                    idx,
                    item.name,
                    item.details or "",
                    float(item.quantity),
                    supplier_name or "",
                    float(item.price),
                    float(item.cost_price) if item.cost_price is not None else "",
                    float(item.commission),
                    item.currency,
                    float(line_total),
                    float(line_total_with_commission),
                    float(profit),
                ]
            )
        else:
            ws_items.append(
                [
                    idx,
                    item.name,
                    item.details or "",
                    float(item.quantity),
                    supplier_name or "",
                    float(item.price),
                    float(item.commission),
                    item.currency,
                    float(line_total),
                    float(line_total_with_commission),
                ]
            )

    _auto_size(ws_items, len(headers))

    # ——— Лист «Заявки на оплату» ———
    ws_req = wb.create_sheet("Заявки на оплату")
    req_headers = [
        "№",
        "Позиции",
        "Сумма",
        "Оплачено",
        "Остаток",
        "Валюта",
        "Приоритет",
        "Дедлайн",
        "Статус",
        "Реквизиты",
        "Детали платежа",
        "Создано",
    ]
    ws_req.append(req_headers)
    _apply_header_style(ws_req, 1, len(req_headers))

    priority_ru = {"urgent": "Срочно", "normal": "Обычно", "deferred": "Отложено"}

    for idx, req in enumerate(requests, start=1):
        items_names = ", ".join(
            ri.project_item.name if ri.project_item else "—" for ri in req.items
        )
        paid = sum(
            (p.amount for p in req.payments if p.status == "confirmed"),
            start=Decimal("0"),
        )
        remaining = req.total_amount - paid
        status = "Оплачено" if remaining <= 0 else "В работе"
        ws_req.append(
            [
                idx,
                items_names,
                float(req.total_amount),
                float(paid),
                float(remaining),
                req.currency,
                priority_ru.get(req.priority, req.priority),
                req.due_date.isoformat() if req.due_date else "",
                status,
                req.requisites or "",
                req.payment_details or "",
                req.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )
    _auto_size(ws_req, len(req_headers))

    # ——— Лист «Сводка по валютам» ———
    # Клиент видит комиссию, но не видит прибыль
    ws_sum = wb.create_sheet("Сводка по валютам")
    sum_headers = (
        ["Валюта", "Итого", "Оплачено", "Остаток", "Комиссия", "Прибыль"]
        if is_admin
        else ["Валюта", "Итого", "Оплачено", "Остаток", "Комиссия"]
    )
    ws_sum.append(sum_headers)
    _apply_header_style(ws_sum, 1, len(sum_headers))

    for currency, (total, commission, profit) in sorted(summary_items.items()):
        paid = summary_paid.get(currency, Decimal("0"))
        remaining = total - paid
        if is_admin:
            ws_sum.append(
                [currency, float(total), float(paid), float(remaining), float(commission), float(profit)]
            )
        else:
            ws_sum.append(
                [currency, float(total), float(paid), float(remaining), float(commission)]
            )
    _auto_size(ws_sum, len(sum_headers))

    # ── Заголовок над позициями (для контекста) ──
    ws_items.insert_rows(1)
    ws_items["A1"] = f"Проект №{project.project_number} — {project.name}"
    ws_items["A1"].font = Font(bold=True, size=14)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
