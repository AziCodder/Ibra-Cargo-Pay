"""
Сервис импорта номенклатуры проекта из Excel (xlsx).

Предоставляет:
- generate_items_template() — создаёт пустой xlsx-шаблон с заголовками
- parse_items_xlsx() — парсит загруженный xlsx, валидирует каждую строку,
  создаёт ProjectItem и возвращает результат с построчными ошибками.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.project_item import ProjectItem
from app.models.supplier import Supplier

TEMPLATE_HEADERS = [
    "name",  # Обязательно
    "details",  # Опционально
    "quantity",  # Обязательно, > 0
    "price",  # Обязательно, >= 0
    "cost_price",  # Опционально, >= 0; пусто — NULL (итог по цене)
    "currency",  # Обязательно: CNY / USD / RUB
    "commission",  # Опционально, >= 0
    "supplier_full_name",  # Опционально
]

TEMPLATE_LABELS_RU = {
    "name": "Наименование *",
    "details": "Детали",
    "quantity": "Количество *",
    "price": "Цена *",
    "cost_price": "Себестоимость",
    "currency": "Валюта (CNY/USD/RUB) *",
    "commission": "Комиссия, %",
    "supplier_full_name": "Поставщик (ФИО из базы)",
}


def generate_items_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Позиции"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")

    labels = [TEMPLATE_LABELS_RU[h] for h in TEMPLATE_HEADERS]
    ws.append(labels)
    # Вторая строка — технические ключи (для справки / парсинга)
    ws.append(TEMPLATE_HEADERS)
    for col in range(1, len(TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    # Пример строки
    ws.append(
        [
            "Пример позиции",
            "Необязательные детали",
            10,
            100.00,
            80.00,
            "CNY",
            5,
            "",
        ]
    )

    # Ширина столбцов
    for col in range(1, len(TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[chr(64 + col)].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _parse_decimal(value, *, allow_none: bool = False) -> Decimal | None:
    if value is None or value == "":
        if allow_none:
            return None
        raise ValueError("Значение обязательно")
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError) as exc:
        raise ValueError(f"Некорректное число: {value!r}") from exc


async def parse_items_xlsx(
    db: AsyncSession, project_id: int, file_bytes: bytes, *, created_by: int
) -> dict:
    """
    Парсит xlsx-файл с позициями и создаёт записи ProjectItem.

    Первая строка игнорируется (человекочитаемые заголовки), вторая строка —
    технические ключи (соответствуют TEMPLATE_HEADERS), с третьей — данные.
    Если пользователь заполнил только одну шапку с русскими названиями —
    также парсим начиная со второй строки, пытаясь понять колонки по позиции.

    Возвращает {"created": N, "errors": [{"row": int, "message": str}]}
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Не удалось прочитать файл: {exc}") from exc

    ws = wb.active
    if ws is None:
        raise ValueError("Файл не содержит листов")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Файл пустой или не содержит данных")

    # Определяем с какой строки идут данные
    first_row = [str(v).strip() if v is not None else "" for v in rows[0]]
    data_start_idx: int
    if len(rows) >= 2 and all(
        v in TEMPLATE_HEADERS
        for v in (str(x).strip() if x is not None else "" for x in rows[1])
        if v
    ):
        # Вторая строка — технические ключи
        data_start_idx = 2
    else:
        # Только человекочитаемые заголовки в первой строке
        data_start_idx = 1

    # Словарь поставщиков по full_name (lowercase) для быстрого lookup
    suppliers_rows = (await db.execute(select(Supplier))).scalars().all()
    supplier_by_name = {s.full_name.strip().lower(): s.id for s in suppliers_rows}

    created = 0
    max_sort_result = await db.execute(
        select(func.coalesce(func.max(ProjectItem.sort_order), -1)).where(
            ProjectItem.project_id == project_id
        )
    )
    next_sort = int(max_sort_result.scalar_one()) + 1
    errors: list[dict] = []

    for row_idx, raw_row in enumerate(rows[data_start_idx:], start=data_start_idx + 1):
        # Пустая строка — пропускаем
        if not raw_row or all(v is None or (isinstance(v, str) and not v.strip()) for v in raw_row):
            continue

        values = list(raw_row) + [None] * (len(TEMPLATE_HEADERS) - len(raw_row))
        row_dict = dict(zip(TEMPLATE_HEADERS, values))

        try:
            name = (row_dict.get("name") or "").strip() if isinstance(row_dict.get("name"), str) else str(row_dict.get("name") or "").strip()
            if not name:
                raise ValueError("Поле 'name' обязательно")

            quantity = _parse_decimal(row_dict.get("quantity"))
            if quantity is None or quantity <= 0:
                raise ValueError("Количество должно быть > 0")

            price = _parse_decimal(row_dict.get("price"))
            if price is None or price < 0:
                raise ValueError("Цена должна быть >= 0")

            cost_price = _parse_decimal(row_dict.get("cost_price"), allow_none=True)
            if cost_price is not None and cost_price < 0:
                raise ValueError("Себестоимость должна быть >= 0")

            currency = str(row_dict.get("currency") or "").strip().upper()
            if currency not in ("CNY", "USD", "RUB"):
                raise ValueError("Валюта должна быть CNY, USD или RUB")

            commission_raw = row_dict.get("commission")
            commission = _parse_decimal(commission_raw, allow_none=True)
            if commission is None:
                commission = Decimal("0")
            if commission < 0:
                raise ValueError("Комиссия должна быть >= 0")

            supplier_name_raw = row_dict.get("supplier_full_name")
            supplier_id: int | None = None
            if supplier_name_raw and str(supplier_name_raw).strip():
                lookup_key = str(supplier_name_raw).strip().lower()
                supplier_id = supplier_by_name.get(lookup_key)
                if supplier_id is None:
                    # Не нашли — создаём позицию без поставщика, фиксируем предупреждение
                    errors.append(
                        {
                            "row": row_idx,
                            "message": f"Поставщик '{supplier_name_raw}' не найден, позиция создана без поставщика",
                        }
                    )

            details = row_dict.get("details")
            details_str = (
                str(details).strip() if details and str(details).strip() else None
            )

            item = ProjectItem(
                project_id=project_id,
                name=name,
                details=details_str,
                quantity=quantity,
                supplier_id=supplier_id,
                price=price,
                cost_price=cost_price,
                currency=currency,
                commission=commission,
                created_by=created_by,
                shared_access=False,
                sort_order=next_sort,
            )
            next_sort += 1
            db.add(item)
            created += 1

        except Exception as exc:  # noqa: BLE001
            errors.append({"row": row_idx, "message": str(exc)})

    return {"created": created, "errors": errors}
